import io
from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

from app.excel_io import ExcelImportError, plan_from_xlsx, plan_to_xlsx
from app.models import Plan, Task


def make_xlsx(rows, header=("задача", "описание", "исполнитель", "длительность", "предшественники")):
    wb = Workbook()
    ws = wb.active
    ws.append(list(header))
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_imports_russian_headers_and_resolves_predecessors_by_name():
    data = make_xlsx(
        [
            ("Собрать требования", "интервью", "Иванов", 5, ""),
            ("Прототип", "фигма", "Петрова", 3, "Собрать требования"),
            ("Тесты", "", "Сидоров", 2, "Прототип, Собрать требования"),
        ]
    )
    plan = plan_from_xlsx(data, project_start=date(2026, 4, 1))

    assert [t.name for t in plan.tasks] == ["Собрать требования", "Прототип", "Тесты"]
    proto = plan.tasks[1]
    assert proto.predecessors == [plan.tasks[0].id]
    assert proto.assignee == "Петрова"
    assert proto.duration_days == 3
    assert set(plan.tasks[2].predecessors) == {plan.tasks[0].id, plan.tasks[1].id}


def test_accepts_english_headers_in_any_order_and_ignores_extra_columns():
    data = make_xlsx(
        [("B", "", "Bob", 2, "A"), ("A", "", "Ann", 1, "")],
        header=("Predecessors", "Task", "Duration", "Assignee", "Description", "Start"),
    )
    # header order above maps: predecessors, task, duration, assignee, description, start
    plan = plan_from_xlsx(
        make_xlsx(
            [("", "A", 1, "Ann", "", "2026-01-01"), ("A", "B", 2, "Bob", "", "")],
            header=("Predecessors", "Task", "Duration", "Assignee", "Description", "Start"),
        ),
        project_start=date(2026, 1, 1),
    )
    assert [t.name for t in plan.tasks] == ["A", "B"]
    assert plan.tasks[1].predecessors == [plan.tasks[0].id]
    assert data  # first workbook unused beyond header exercise


def test_predecessors_may_reference_ids_or_semicolons():
    data = make_xlsx([("A", "", "", 1, ""), ("B", "", "", 1, "a;A")])
    plan = plan_from_xlsx(data, project_start=date(2026, 1, 1))
    assert plan.tasks[1].predecessors == [plan.tasks[0].id]


def test_unknown_predecessor_reports_row_number():
    data = make_xlsx([("A", "", "", 1, "Ghost")])
    with pytest.raises(ExcelImportError) as err:
        plan_from_xlsx(data, project_start=date(2026, 1, 1))
    assert "2" in err.value.details[0]  # header is row 1, so the task is row 2
    assert "Ghost" in err.value.details[0]


def test_bad_duration_reports_row_number():
    data = make_xlsx([("A", "", "", "две недели", "")])
    with pytest.raises(ExcelImportError) as err:
        plan_from_xlsx(data, project_start=date(2026, 1, 1))
    assert "2" in err.value.details[0]


def test_missing_task_column_is_rejected():
    data = make_xlsx([("A", 1)], header=("описание", "длительность"))
    with pytest.raises(ExcelImportError) as err:
        plan_from_xlsx(data, project_start=date(2026, 1, 1))
    assert "задача" in str(err.value).lower() or "task" in str(err.value).lower()


def test_duplicate_names_get_distinct_ids():
    data = make_xlsx([("Ревью", "", "", 1, ""), ("Ревью", "", "", 2, "")])
    plan = plan_from_xlsx(data, project_start=date(2026, 1, 1))
    assert plan.tasks[0].id != plan.tasks[1].id


def test_blank_rows_are_skipped():
    data = make_xlsx([("A", "", "", 1, ""), (None, None, None, None, None), ("B", "", "", 1, "A")])
    plan = plan_from_xlsx(data, project_start=date(2026, 1, 1))
    assert len(plan.tasks) == 2


def test_export_round_trips_through_import():
    plan = Plan(
        project_start=date(2026, 5, 4),
        tasks=[
            Task(id="t1", name="Анализ", description="d1", assignee="Иванов", duration_days=4),
            Task(id="t2", name="Дизайн", assignee="Петрова", duration_days=2, predecessors=["t1"]),
            Task(
                id="t3",
                name="Сдача",
                duration_days=1,
                predecessors=["t2"],
                start_no_earlier_than=date(2026, 6, 1),
            ),
        ],
    )
    blob = plan_to_xlsx(plan)
    again = plan_from_xlsx(blob, project_start=date(2026, 5, 4))

    assert [t.name for t in again.tasks] == ["Анализ", "Дизайн", "Сдача"]
    assert again.tasks[1].predecessors == [again.tasks[0].id]
    assert again.tasks[0].assignee == "Иванов"
    assert again.tasks[0].duration_days == 4
    assert again.project_start == date(2026, 5, 4)


def test_export_includes_computed_dates():
    plan = Plan(
        project_start=date(2026, 5, 4),
        tasks=[Task(id="t1", name="Анализ", duration_days=4)],
    )
    ws = load_workbook(io.BytesIO(plan_to_xlsx(plan))).active
    header = [c.value for c in ws[1]]
    assert "начало" in [str(h).lower() for h in header]
    row = [c.value for c in ws[2]]
    assert "2026-05-04" in [str(v) for v in row]


def test_export_keeps_ids_stable_through_a_round_trip():
    plan = Plan(
        project_start=date(2026, 5, 4),
        tasks=[
            Task(id="analiz", name="Анализ", duration_days=2),
            # a name containing the separator used for predecessor lists
            Task(id="ct-phase1", name="Исследование, фаза I", duration_days=3, predecessors=["analiz"]),
        ],
    )
    again = plan_from_xlsx(plan_to_xlsx(plan), project_start=date(2026, 5, 4))
    assert [t.id for t in again.tasks] == ["analiz", "ct-phase1"]
    assert again.tasks[1].predecessors == ["analiz"]


def test_duplicate_ids_in_a_file_are_made_unique():
    data = make_xlsx(
        [("dup", "A", "", "", 1, ""), ("dup", "B", "", "", 2, "")],
        header=("id", "задача", "описание", "исполнитель", "длительность", "предшественники"),
    )
    plan = plan_from_xlsx(data, project_start=date(2026, 1, 1))
    assert [t.id for t in plan.tasks] == ["dup", "dup-2"]


def test_status_round_trips_through_excel():
    plan = Plan(
        project_start=date(2026, 5, 4),
        tasks=[
            Task(id="a", name="Анализ", duration_days=2, status="done", progress=100),
            Task(id="b", name="Дизайн", duration_days=2, status="blocked", predecessors=["a"]),
        ],
    )
    again = plan_from_xlsx(plan_to_xlsx(plan), project_start=date(2026, 5, 4))
    assert [t.status.value for t in again.tasks] == ["done", "blocked"]


def test_status_column_accepts_labels_and_codes():
    data = make_xlsx(
        [("A", "", "", 1, "", "в работе"), ("B", "", "", 1, "", "done"), ("C", "", "", 1, "", "")],
        header=("задача", "описание", "исполнитель", "длительность", "предшественники", "статус"),
    )
    plan = plan_from_xlsx(data, project_start=date(2026, 1, 1))
    assert [t.status.value for t in plan.tasks] == ["in_progress", "done", "planned"]


def test_unknown_status_in_file_is_reported_with_the_row():
    data = make_xlsx(
        [("A", "", "", 1, "", "почти готово")],
        header=("задача", "описание", "исполнитель", "длительность", "предшественники", "статус"),
    )
    with pytest.raises(ExcelImportError) as err:
        plan_from_xlsx(data, project_start=date(2026, 1, 1))
    assert "строка 2" in err.value.details[0]
