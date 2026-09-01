"""Excel import / export.

Input format required by the task: columns `задача, описание, исполнитель,
длительность, предшественники`. Real spreadsheets are messy, so the importer is
deliberately tolerant: header names are matched against Russian and English
aliases in any order, extra columns are ignored, blank rows are skipped, and
predecessors may reference either task names or ids, separated by commas or
semicolons.

Errors are collected per row and reported together — a user fixing their file
wants the whole list, not the first failure.
"""

from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime
from typing import Any, Optional

from openpyxl import Workbook, load_workbook

from .models import STATUS_LABELS, Plan, Task, TaskStatus

# Columns written on export. The first five are the required input format;
# the rest are computed and ignored when the file is imported back.
EXPORT_HEADER = [
    "id",
    "задача",
    "описание",
    "исполнитель",
    "длительность",
    "предшественники",
    "зафиксировать с",
    "начало",
    "окончание",
    "статус",
    "критический путь",
    "запас, дн",
    "прогресс, %",
]

_ALIASES: dict[str, set[str]] = {
    "id": {"id", "ид", "идентификатор", "код", "key"},
    "name": {"задача", "задачи", "название", "наименование", "работа", "task", "task name", "name"},
    "description": {"описание", "комментарий", "детали", "description", "details", "notes"},
    "assignee": {"исполнитель", "ответственный", "ресурс", "assignee", "owner", "resource"},
    "duration": {
        "длительность",
        "длительность, дн",
        "продолжительность",
        "дни",
        "duration",
        "duration (days)",
        "days",
    },
    "predecessors": {
        "предшественники",
        "предшественник",
        "зависимости",
        "зависит от",
        "predecessors",
        "predecessor",
        "depends on",
        "dependencies",
        "deps",
    },
    "pin": {"зафиксировать с", "фиксация", "не раньше", "pin", "start no earlier than"},
    "status": {"статус", "состояние", "status", "state"},
}


class ExcelImportError(Exception):
    """Raised when a spreadsheet cannot be turned into a plan."""

    def __init__(self, message: str, details: Optional[list[str]] = None) -> None:
        super().__init__(message)
        self.message = message
        self.details = details or []


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower()).rstrip(":.")


def _map_columns(header_row: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        key = _norm_header(raw)
        if not key:
            continue
        for field, aliases in _ALIASES.items():
            if field not in mapping and key in aliases:
                mapping[field] = idx
                break
    if "name" not in mapping:
        raise ExcelImportError(
            "В файле не найдена колонка «задача»",
            [
                "Ожидаются колонки: задача, описание, исполнитель, длительность, предшественники. "
                f"Найдено: {', '.join(str(h) for h in header_row if h) or '—'}"
            ],
        )
    return mapping


def unique_id(raw: str, taken: set[str]) -> str:
    """Keep an id supplied in the file, making it unique if the file repeats it."""
    candidate = re.sub(r"[^A-Za-z0-9_-]+", "-", raw.strip()).strip("-")[:40]
    if not candidate:
        return ""
    base, n = candidate, 2
    while candidate in taken:
        candidate = f"{base}-{n}"
        n += 1
    taken.add(candidate)
    return candidate


def slugify(name: str, taken: set[str]) -> str:
    """Human-readable stable id: `sobrat-trebovaniya`, `sobrat-trebovaniya-2`...

    The agent sees these ids in tool calls, so readable ids make its reasoning
    (and our debugging) far easier than opaque UUIDs.
    """
    ascii_name = unicodedata.normalize("NFKD", _translit(name.lower()))
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_name).strip("-")[:40] or "task"
    candidate = slug
    n = 2
    while candidate in taken:
        candidate = f"{slug}-{n}"
        n += 1
    taken.add(candidate)
    return candidate


_TRANSLIT = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "zh",
    "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o",
    "п": "p", "р": "r", "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "ts",
    "ч": "ch", "ш": "sh", "щ": "sch", "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu",
    "я": "ya",
}


def _translit(text: str) -> str:
    return "".join(_TRANSLIT.get(ch, ch) for ch in text)


def _cell(row: tuple[Any, ...], idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _as_duration(value: Any, row_no: int, errors: list[str]) -> int:
    if value is None or _as_text(value) == "":
        return 1
    try:
        number = float(_as_text(value).replace(",", "."))
    except ValueError:
        errors.append(
            f"строка {row_no}: длительность «{_as_text(value)}» — не число"
        )
        return 1
    days = int(round(number))
    if days < 1:
        errors.append(f"строка {row_no}: длительность должна быть не меньше 1 дня")
        return 1
    return days


def _as_date(value: Any, row_no: int, errors: list[str]) -> Optional[date]:
    if value is None or _as_text(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _as_text(value)
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    errors.append(f"строка {row_no}: дату «{text}» не удалось разобрать (ожидается ГГГГ-ММ-ДД)")
    return None


def _as_status(value: Any, row_no: int, errors: list[str]) -> TaskStatus:
    text = _as_text(value)
    if not text:
        return TaskStatus.PLANNED
    from .ops import OpError, parse_status  # локальный импорт: ops зависит от нас

    try:
        return parse_status(text)
    except OpError as exc:
        errors.append(f"строка {row_no}: {exc}")
        return TaskStatus.PLANNED


def _split_predecessors(value: Any) -> list[str]:
    text = _as_text(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"[;,\n]", text) if part.strip()]


def plan_from_xlsx(blob: bytes, project_start: date, title: Optional[str] = None) -> Plan:
    """Parse an .xlsx file into a `Plan`.

    Raises `ExcelImportError` with a per-row breakdown if anything is wrong.
    """
    try:
        wb = load_workbook(io.BytesIO(blob), data_only=True)
    except Exception as exc:  # openpyxl raises a zoo of exceptions
        raise ExcelImportError("Не удалось прочитать файл как .xlsx", [str(exc)]) from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ExcelImportError("Файл пустой")

    columns = _map_columns(list(rows[0]))
    errors: list[str] = []
    taken_ids: set[str] = set()
    raw_rows: list[tuple[int, Task, list[str]]] = []

    for offset, row in enumerate(rows[1:], start=2):
        if row is None or all(_as_text(c) == "" for c in row):
            continue
        name = _as_text(_cell(row, columns.get("name")))
        if not name:
            errors.append(f"строка {offset}: не заполнено название задачи")
            continue
        supplied_id = unique_id(_as_text(_cell(row, columns.get("id"))), taken_ids)
        task = Task(
            id=supplied_id or slugify(name, taken_ids),
            name=name,
            description=_as_text(_cell(row, columns.get("description"))),
            assignee=_as_text(_cell(row, columns.get("assignee"))),
            duration_days=_as_duration(_cell(row, columns.get("duration")), offset, errors),
            predecessors=[],
            start_no_earlier_than=_as_date(_cell(row, columns.get("pin")), offset, errors),
            status=_as_status(_cell(row, columns.get("status")), offset, errors),
        )
        raw_rows.append((offset, task, _split_predecessors(_cell(row, columns.get("predecessors")))))

    by_name = {t.name.strip().lower(): t.id for _, t, _ in raw_rows}
    by_id = {t.id: t.id for _, t, _ in raw_rows}

    for row_no, task, refs in raw_rows:
        resolved: list[str] = []
        for ref in refs:
            target = by_name.get(ref.lower()) or by_id.get(ref) or by_id.get(ref.lower())
            if target is None:
                errors.append(
                    f"строка {row_no}: предшественник «{ref}» не найден среди задач файла"
                )
            elif target == task.id:
                errors.append(f"строка {row_no}: задача не может зависеть от себя")
            elif target not in resolved:
                resolved.append(target)
        task.predecessors = resolved

    if errors:
        raise ExcelImportError("Файл не прошёл проверку", errors)

    return Plan(
        title=title or "План проекта",
        project_start=project_start,
        tasks=[t for _, t, _ in raw_rows],
    )


def _predecessor_cell(task_ids: list[str], names: dict[str, str]) -> str:
    """Render a predecessor list so that re-importing the file gives the same graph.

    Names are friendlier for a human editing the sheet, but the separator is a
    comma — so a name that itself contains a comma or semicolon is written as
    the task id instead (ids never contain either).
    """
    rendered = []
    for task_id in task_ids:
        name = names.get(task_id, task_id)
        rendered.append(task_id if re.search(r"[;,\n]", name) else name)
    return ", ".join(rendered)


def plan_to_xlsx(plan: Plan) -> bytes:
    """Serialise a plan to .xlsx: the input columns plus computed schedule data."""
    from .scheduler import schedule_plan  # local import avoids a cycle

    schedule = schedule_plan(plan)
    computed = {t.id: t for t in schedule.tasks}
    names = {t.id: t.name for t in plan.tasks}

    wb = Workbook()
    ws = wb.active
    ws.title = "План"
    ws.append(EXPORT_HEADER)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    for task in plan.tasks:
        c = computed[task.id]
        ws.append(
            [
                task.id,
                task.name,
                task.description,
                task.assignee,
                task.duration_days,
                _predecessor_cell(task.predecessors, names),
                task.start_no_earlier_than.isoformat() if task.start_no_earlier_than else "",
                c.start.isoformat(),
                c.end.isoformat(),
                STATUS_LABELS[task.status],
                "да" if c.is_critical else "",
                c.slack_days,
                task.progress,
            ]
        )

    widths = [22, 38, 46, 18, 14, 34, 18, 14, 14, 18, 18, 12, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = "A2"

    meta = wb.create_sheet("Метаданные")
    meta.append(["название проекта", plan.title])
    meta.append(["старт проекта", plan.project_start.isoformat()])
    meta.append(["окончание (расчёт)", schedule.project_end.isoformat() if schedule.project_end else ""])
    meta.column_dimensions["A"].width = 26
    meta.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def meta_from_xlsx(blob: bytes) -> tuple[Optional[str], Optional[date]]:
    """Read title and project start back from a file we exported earlier.

    Files from other tools simply have no metadata sheet, and both values fall
    back to caller-supplied defaults.
    """
    try:
        wb = load_workbook(io.BytesIO(blob), data_only=True)
    except Exception:
        return None, None
    if "Метаданные" not in wb.sheetnames:
        return None, None

    title: Optional[str] = None
    start: Optional[date] = None
    for key, value in wb["Метаданные"].iter_rows(values_only=True):
        field = _norm_header(key)
        if field == "название проекта" and _as_text(value):
            title = _as_text(value)
        elif field == "старт проекта":
            start = _as_date(value, 0, [])
    return title, start
